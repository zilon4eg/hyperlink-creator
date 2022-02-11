import xlwings
from openpyxl import load_workbook


class Excel:
    def __init__(self, registry_path, dir_scan, ws_name, settings):
        self.settings = settings
        self.font_size = round(int(settings['font']['size']), 1)
        self.font_name = settings['font']['name']
        self.hyperlink_color = settings['font']['color']
        self.dir_scan = dir_scan

        if registry_path not in [True, None, 'None', '']:
            self.wb = xlwings.Book(registry_path)
        else:
            self.wb = xlwings.books.active

        if ws_name not in [True, None, 'None', '']:
            self.ws_name = ws_name
        else:
            self.ws_name = self.wb.sheets.active.name

        print(f'Документ "{self.wb.name}" открыт.')
        if self.ws_name not in list(sheet.name for sheet in self.wb.sheets):
            print(f'Лист "{self.ws_name}" отсутствует в книге.')
        self.ws = self.wb.sheets[self.ws_name]
        print(f'Выбран лист "{self.ws_name}".')

    def size_column(self, column):
        column = Excel.number_to_letter(column) if type(column) is int else column
        last_cell_with_data = None
        last_cell = 1
        empty_cell_count = 0

        while True:
            cell_data = self.ws[f'{column}{last_cell}'].value
            if cell_data in [None, '']:
                empty_cell_count += 1
                last_cell += 1
            else:
                empty_cell_count = 0
                last_cell_with_data = last_cell
                last_cell += 1
            if empty_cell_count > 30:
                return last_cell_with_data

    def size_string(self, string):
        last_cell_with_data = None
        last_cell = 1
        empty_cell_count = 0

        while True:
            cell_data = self.ws[f'{Excel.number_to_letter(last_cell)}{string}'].value
            if cell_data in [None, '']:
                empty_cell_count += 1
                last_cell += 1
            else:
                empty_cell_count = 0
                last_cell_with_data = last_cell
                last_cell += 1
            if empty_cell_count > 30:
                return last_cell_with_data

    def get_path_active_book(self):
        return self.wb.fullname

    def create_hyperlinks(self, name, file_name, position):
        # ===(Формируем список параметров текста)===
        is_bold = self.settings['font']['style']['bold']
        is_italic = self.settings['font']['style']['italic']
        is_underline = self.settings['font']['style']['underline']
        # ===(Переменная для определения, была ли пересоздана гиперссылка)===
        hyperlink_is_ceate = False

        try:
            hyperlink = self.ws[f'{position}'].hyperlink
        except Exception:
            hyperlink = None

        if (
                name == self.ws[f'{position}'].value
                and f'{self.dir_scan}\\{file_name}'.replace('\\', '/') in hyperlink.replace('\\', '/')
        ):
            hyperlink_is_ceate = False
        else:
            self.ws[f'{position}'].add_hyperlink(f'{self.dir_scan}\\{file_name}', name)
            hyperlink_is_ceate = True

        if hyperlink_is_ceate is False:
            if self.font_name != self.ws[f'{position}'].font.name:
                self.ws[f'{position}'].font.name = self.font_name
            if self.font_size != self.ws[f'{position}'].font.size:
                self.ws[f'{position}'].font.size = self.font_size
            if Excel.hex_to_rgb(self.hyperlink_color) != self.ws[f'{position}'].font.color:
                self.ws[f'{position}'].font.color = self.hyperlink_color
            if is_bold != self.ws[f'{position}'].font.bold:
                self.ws[f'{position}'].font.bold = is_bold
            if is_italic != self.ws[f'{position}'].font.italic:
                self.ws[f'{position}'].font.italic = is_italic
            if is_underline != self.ws[f'{position}'].api.Font.Underline:
                self.ws[f'{position}'].api.Font.Underline = is_underline  # True == 2 - single, 3 - double
            self.ws[f'{position}'].api.HorizontalAlignment = -4108
            self.ws[f'{position}'].api.VerticalAlignment = -4108
            # self.borders_all(f'{position}')
        else:
            self.ws[f'{position}'].font.name = self.font_name
            self.ws[f'{position}'].font.size = self.font_size
            self.ws[f'{position}'].font.color = self.hyperlink_color
            self.ws[f'{position}'].font.bold = is_bold
            self.ws[f'{position}'].font.italic = is_italic
            self.ws[f'{position}'].api.Font.Underline = is_underline  # True == 2 - single, 3 - double
            self.ws[f'{position}'].api.HorizontalAlignment = -4108
            self.ws[f'{position}'].api.VerticalAlignment = -4108
            self.borders_all(f'{position}')

    def borders_all(self, cell):
        for i in range(7, 13):
            self.ws[cell].api.Borders(i).LineStyle = 1

    @staticmethod
    def hex_to_rgb(hex_color):
        h = hex_color.lstrip('#')
        return tuple(int(h[i:i + 2], 16) for i in (0, 2, 4))

    @staticmethod
    def number_to_letter(number):
        return chr(int(((number - 1) / 26) + 64)) + chr(int(((number - 1) % 26) + 1 + 64)) if number > 26 else chr(
            int(number + 64))


def get_all_ws(file_path):
    wb = load_workbook(file_path)
    ws_list = wb.sheetnames
    return ws_list


if __name__ == '__main__':
    pass
